# apps/gso_reports/views.py
import os
import csv
import json
import calendar
import openpyxl
from datetime import datetime
from django.utils import timezone
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import range_boundaries


from apps.gso_requests.models import ServiceRequest, Feedback
from apps.gso_accounts.models import User, Unit
from .models import WorkAccomplishmentReport, SuccessIndicator, IPMT
from .utils import normalize_report
from apps.ai_service.utils import generate_war_description
from apps.ai_service.utils import generate_ipmt_summary_sync


# -------------------------------
# Role Checks
# -------------------------------
def is_gso_or_director(user):
    return user.is_authenticated and user.role in ["gso", "director"]


# -------------------------------
# Accomplishment Report View
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def accomplishment_report(request):
    completed_requests = ServiceRequest.objects.filter(status="Completed").order_by("-created_at")
    all_wars = WorkAccomplishmentReport.objects.select_related("request", "unit") \
        .prefetch_related("assigned_personnel").all().order_by("-date_started")

    reports = []
    war_request_ids = set(war.request_id for war in all_wars if war.request_id)

    # Process completed requests (no WAR yet)
    for r in completed_requests:
        if r.id in war_request_ids:
            continue
        norm = normalize_report(r)
        norm["id"] = r.id

        if not norm.get("description") or not norm["description"].strip():
            try:
                desc = generate_war_description(
                    activity_name=getattr(r, "activity_name", getattr(r, "title", "Task")),
                    unit=getattr(r.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in r.assigned_personnel.all()] if hasattr(r, "assigned_personnel") else None
                )
                r.description = desc or "No description generated."
                r.save(update_fields=["description"])
                norm["description"] = r.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Process existing WARs
    for war in all_wars:
        norm = normalize_report(war)
        norm["id"] = war.id

        if not norm.get("description") or not norm["description"].strip():
            try:
                desc = generate_war_description(
                    activity_name=getattr(war, "activity_name", getattr(war, "title", "Task")),
                    unit=getattr(war.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in war.assigned_personnel.all()] if hasattr(war, "assigned_personnel") else None
                )
                war.description = desc or "No description generated."
                war.save(update_fields=["description"])
                norm["description"] = war.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Filters
    search_query = request.GET.get("q")
    if search_query:
        reports = [r for r in reports if search_query.lower() in str(r).lower()]

    unit_filter = request.GET.get("unit")
    if unit_filter:
        reports = [r for r in reports if r["unit"].lower() == unit_filter.lower()]

    reports.sort(key=lambda r: r["date"], reverse=True)

    personnel_qs = User.objects.filter(role="personnel", account_status="active") \
        .select_related('unit').order_by('unit__name', 'first_name')

    personnel_list = [
        {
            "full_name": u.get_full_name() or u.username,
            "username": u.username,
            "unit": u.unit.name.lower() if u.unit else "unassigned"
        }
        for u in personnel_qs
    ]

    return render(
        request,
        "gso_office/accomplishment_report/accomplishment_report.html",
        {"reports": reports, "personnel_list": personnel_list},
    )

@login_required
@user_passes_test(is_gso_or_director)
@csrf_exempt  # allows AJAX POST from JS
def update_success_indicator(request):
    """
    Updates the Success Indicator of a Work Accomplishment Report via AJAX.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            war_id = data.get("war_id")
            indicator_id = data.get("indicator_id")

            war = WorkAccomplishmentReport.objects.get(id=war_id)
            indicator = SuccessIndicator.objects.get(id=indicator_id)

            war.success_indicator = indicator
            war.save(update_fields=["success_indicator"])

            return JsonResponse({"success": True, "indicator": indicator.code})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})

# -------------------------------
# Generate IPMT Excel
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def generate_ipmt(request):
    reports = []
    personnel_list = []

    if request.method == "POST":
        try:
            body = json.loads(request.body.decode("utf-8"))
            month_filter = body.get("month")
            unit_filter = body.get("unit")
            personnel_param = body.get("personnel", "")
            reports = body.get("rows", [])
        except Exception:
            month_filter = request.POST.get("month")
            unit_filter = request.POST.get("unit")
            personnel_param = request.POST.get("personnel", "")
            rows_data = request.POST.get("rows", "[]")
            try:
                reports = json.loads(rows_data)
            except json.JSONDecodeError:
                reports = []

        personnel_list = [p.strip() for p in personnel_param.split(",") if p.strip()]

        # Clean indicators
        for r in reports:
            if not r.get("indicator"):
                continue
            code_only = r["indicator"].split(" - ")[0].strip()
            si = SuccessIndicator.objects.filter(code__iexact=code_only).first()
            if si:
                r["indicator"] = f"{si.code} - {si.description}"

    else:
        return HttpResponse("Only POST allowed.", status=400)

    template_path = os.path.join(settings.BASE_DIR, "static", "excel_file", "sampleipmt.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Only support single personnel for header fields
    if personnel_list:
        user_obj = get_user_by_identifier(personnel_list[0])
    else:
        user_obj = None

    # --- Header Section ---
    ws["B7"] = user_obj.unit.name if user_obj and user_obj.unit else "No Unit"
    ws["B8"] = user_obj.get_full_name() if user_obj else "No Name"
    ws["B9"] = user_obj.employment_status.employment_status if user_obj and user_obj.employment_status else "No Status"
    ws["B10"] = user_obj.position.name if user_obj and user_obj.position else "No Position"

    # Month
    if "-" in month_filter:
        year, month_num = map(int, month_filter.split("-"))
        month_name = f"{calendar.month_name[month_num]} {year}"
    else:
        month_name = month_filter
    ws["B11"] = month_name

    # Define a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --- Data Table ---
    start_row = 13
    for i, r in enumerate(reports, start=start_row):
        # Clean indicator string
        indicator_raw = r.get("indicator", "")
        indicator_raw = " ".join(indicator_raw.split())
        if " - " in indicator_raw:
            code, desc = indicator_raw.split(" - ", 1)
            code = code.strip()
            desc = " ".join(desc.split())
            indicator_clean = f"{code} - {desc}"
        else:
            indicator_clean = indicator_raw.strip()

        desc_clean = " ".join((r.get("description", "") or "").split())
        remarks_clean = " ".join((r.get("remarks", "") or "").split())

        # Write to Excel
        ws.cell(row=i, column=1).value = indicator_clean
        ws.cell(row=i, column=2).value = desc_clean
        ws.cell(row=i, column=3).value = remarks_clean

        # Merge columns C and D for this row
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)

        # Center alignment for all columns A-D
        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=i, column=col).border = thin_border

    thin_side = Side(border_style="thin", color="000000")

    def outside_border(top=False, left=False, bottom=False, right=False):
        return Border(
            top=thin_side if top else None,
            left=thin_side if left else None,
            bottom=thin_side if bottom else None,
            right=thin_side if right else None
        )

    footer_start = start_row + len(reports)  # after last data row

    # --- IPCR note ---
    ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=4)
    note_cell = ws.cell(row=footer_start, column=1)
    note_cell.value = ("(*Based on the IPCR Major Final Output (MFO)/ Program, Activity and Project (PAP), "
                    "select only those success indicators where the accomplishments for the period are aligned to*)")
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[footer_start].height = 40

    # Outside border for IPCR note
    for col in range(1, 5):
        ws.cell(row=footer_start, column=col).border = outside_border(
            top=True, left=(col==1), bottom=True, right=(col==4)
        )

    # --- Prepared by (left box) ---
    prepared_row_start = footer_start + 1
    for i, text in enumerate(["Prepared by:", ws["B8"].value if user_obj else "No Name", "Employee"]):
        row = prepared_row_start + i
        ws.cell(row=row, column=1).value = text
        if i == 1:
            ws.cell(row=row, column=1).font = Font(underline='single')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        # Left box border: top, left, bottom only, no right
        ws.cell(row=row, column=1).border = outside_border(
            top=(i==0),
            left=True,
            bottom=(i==2),
            right=False
        )

    # Add left border, top/bottom, and bottom across B as well
    if i == 2:  # last row of box
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = outside_border(top=False, left=(col==1), bottom=True, right=False)
    else:
        ws.cell(row=row, column=1).border = outside_border(top=(i==0), left=True, bottom=False, right=False)

    # --- Checked and Verified by (right box) ---
    director = User.objects.filter(role="director").first()
    checked_texts = [
        "Checked and Verified by:",
        director.get_full_name() if director else "Director Name",
        "(Department Head / Supervisor)"
    ]

    for i, text in enumerate(checked_texts):
        row = prepared_row_start + i
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3).value = text
        if i == 1:
            ws.cell(row=row, column=3).font = Font(underline='single')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        # Right box border: top, right, bottom only, no left
        for col in range(3, 5):
            ws.cell(row=row, column=col).border = outside_border(
                top=(i==0),
                left=False,
                bottom=(i==2),
                right=True
            )

    # --- Save Response ---
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"IPMT_{unit_filter}_{month_filter}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



# -------------------------------
# Helper
# -------------------------------
def get_user_by_identifier(identifier):
    identifier = identifier.strip()
    if not identifier:
        return None

    user = User.objects.filter(username__iexact=identifier).first()
    if user:
        return user

    parts = identifier.split()
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        user = User.objects.filter(Q(first_name__iexact=first) & Q(last_name__iexact=last)).first()
        if user:
            return user

    return User.objects.filter(Q(first_name__icontains=identifier) | Q(last_name__icontains=identifier)).first()

# -------------------------------
# Get WAR Description (AJAX)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def get_war_description(request, war_id):
    try:
        war = WorkAccomplishmentReport.objects.get(id=war_id)
        return JsonResponse({'description': war.description or ""})
    except WorkAccomplishmentReport.DoesNotExist:
        return JsonResponse({'error': 'WAR not found'}, status=404)
    

# -------------------------------
# Preview IPMT (Web) with AI Summarization
# -------------------------------

@login_required
@user_passes_test(is_gso_or_director)
def preview_ipmt(request):
    """
    Preview IPMT rows for the selected unit, personnel, and month.
    Fetches WARs linked to each SuccessIndicator and assigned personnel.
    AI summarizes multiple WARs per Success Indicator into one concise description.
    """
    month_filter = request.GET.get("month")
    unit_filter = request.GET.get("unit")
    personnel_names = request.GET.getlist("personnel[]") or []

    if not month_filter:
        return HttpResponse("Month is required in 'YYYY-MM' format.", status=400)

    try:
        year, month_num = map(int, month_filter.split("-"))
    except ValueError:
        return HttpResponse("Invalid month format. Use YYYY-MM.", status=400)

    unit = Unit.objects.filter(name__iexact=unit_filter).first()
    if not unit:
        return HttpResponse("Unit not found.", status=404)

    reports = []

    for person_name in personnel_names:
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        # Fetch all active success indicators under the unit
        indicators = SuccessIndicator.objects.filter(unit=unit, is_active=True)

        for indicator in indicators:
            # Get WARs for this user and indicator within the selected month
            wars = WorkAccomplishmentReport.objects.filter(
                unit=unit,
                assigned_personnel=user,
                success_indicator=indicator,
                date_started__year=year,
                date_started__month=month_num
            )

            war_ids = [w.id for w in wars]

            # --- AI summarization per Success Indicator ---
            ipmt_obj = IPMT.objects.filter(
                personnel=user,
                unit=unit,
                month=month_filter,
                indicator=indicator
            ).first()

            if ipmt_obj and ipmt_obj.accomplishment:
                # Use saved summary
                summary_text = ipmt_obj.accomplishment
            else:
                # First-time AI generation
                if war_ids:
                    if len(war_ids) == 1:
                        # Get the single WAR directly — NO AI summarization
                        single_war = WorkAccomplishmentReport.objects.get(id=war_ids[0])
                        summary_text = (
                            getattr(single_war, "work_done", None)
                            or getattr(single_war, "description", None)
                            or ""
                        )

                    else:
                        # AI summarization for multiple WARs
                        summary_dict = generate_ipmt_summary_sync(war_ids)
                        summary_text = summary_dict.get(indicator.code, "")
                else:
                    summary_text = ""

            reports.append({
                "indicator": indicator.code,
                "indicator_description": indicator.description,
                "description": summary_text,  # AI summarized description
                "remarks": "COMPLIED" if summary_text else "",
                "war_ids": war_ids,
            })

    context = {
        "reports": reports,
        "month_filter": month_filter,
        "unit_filter": unit_filter,
        "personnel_names": personnel_names,
    }

    return render(request, "gso_office/reports/ipmt_preview.html", context)



# -------------------------------
# Save IPMT (Web)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def save_ipmt(request):
    """
    Save or update IPMT rows after preview.
    Handles edited descriptions and remarks, links WARs to each row.
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    try:
        data = json.loads(request.body)
        month = data.get("month")
        unit_name = data.get("unit")
        personnel_names = data.get("personnel", [])
        rows = data.get("rows", [])
    except Exception as e:
        return JsonResponse({"error": f"Invalid JSON: {str(e)}"}, status=400)

    unit = Unit.objects.filter(name__iexact=unit_name).first()
    if not unit:
        return JsonResponse({"error": "Unit not found"}, status=404)

    for person_name in personnel_names:
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        for row in rows:
            indicator_code = row.get("indicator")
            if not indicator_code:
                continue

            # Ensure SuccessIndicator exists
            si = SuccessIndicator.objects.filter(unit=unit, code__iexact=indicator_code).first()
            if not si:
                si = SuccessIndicator.objects.create(
                    unit=unit,
                    code=indicator_code,
                    description=row.get("description", ""),
                    is_active=True
                )

            # Get WARs linked to this row
            war_ids = row.get("war_ids", [])
            wars = WorkAccomplishmentReport.objects.filter(assigned_personnel=user, unit=unit)
            if war_ids:
                wars = wars.filter(id__in=war_ids)

            # Take description and remarks from frontend
            accomplishment = row.get("description", "").strip()
            remarks = row.get("remarks", "").strip() or accomplishment

            # Save or update IPMT object
            ipmt_obj, _ = IPMT.objects.update_or_create(
                personnel=user,
                unit=unit,
                month=month,
                indicator=si,
                defaults={
                    "accomplishment": accomplishment,
                    "remarks": remarks
                }
            )

            # Link WARs to IPMT
            ipmt_obj.reports.set(wars)

    return JsonResponse({"status": "success"})

# -------------------------------
# GSO OFFICE: Edit Success Indicator in WAR
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def update_war_success_indicator(request, war_id):
    """
    Allow GSO Office to edit the Success Indicator directly from the WAR page.
    """
    war = get_object_or_404(WorkAccomplishmentReport, id=war_id)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()

        if name:
            indicator, _ = SuccessIndicator.objects.get_or_create(name=name)
            if description:
                indicator.description = description
                indicator.save(update_fields=["description"])
            war.success_indicator = indicator
            war.save(update_fields=["success_indicator"])
            messages.success(request, "Success Indicator updated for WAR.")
            return redirect("gso_reports:war_detail", war_id=war.id)

        messages.warning(request, "Indicator name is required.")

    return render(request, "gso_office/war_success_indicator_form.html", {
        "war": war,
    })








@login_required
@user_passes_test(is_gso_or_director)
def feedback_reports(request):
    """Show feedbacks with optional filtering by unit and date range, and export CSV."""
    
    feedback_list = Feedback.objects.select_related("request", "request__requestor", "request__unit").order_by("-date_submitted")
    units = Unit.objects.all()

    # Get filter parameters from GET
    unit_id = request.GET.get("unit_id")
    month = request.GET.get("month")          # format: 'YYYY-MM'
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Filter by unit
    if unit_id:
        feedback_list = feedback_list.filter(request__unit_id=unit_id)

    # Filter by month
    if month:
        try:
            year, month_num = map(int, month.split("-"))
            feedback_list = feedback_list.filter(
                date_submitted__year=year,
                date_submitted__month=month_num
            )
        except ValueError:
            pass

    # Filter by custom date range
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            feedback_list = feedback_list.filter(date_submitted__date__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            feedback_list = feedback_list.filter(date_submitted__date__lte=end)
        except ValueError:
            pass

    # CSV Export
    if "export" in request.GET:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="feedback_report.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Service Request ID",
            "Unit",
            "Requestor Name",
            "Requestor Email",
            "SQD1","SQD2","SQD3","SQD4","SQD5","SQD6","SQD7","SQD8","SQD9",
            "CC1","CC2","CC3",
            "Average Score",
            "Suggestions",
            "Date Submitted"
        ])
        for fb in feedback_list:
            req = fb.request
            requestor_name = req.custom_full_name or (req.requestor.get_full_name() if req.requestor else "")
            requestor_email = req.custom_email or (req.requestor.email if req.requestor else "")
            request_id = req.id
            unit_name = req.unit.name if req.unit else ""
            formatted_date = fb.date_submitted.strftime("%Y-%m-%d %H:%M") if fb.date_submitted else ""
            writer.writerow([
                request_id,
                unit_name,
                requestor_name,
                requestor_email,
                fb.sqd1 or "", fb.sqd2 or "", fb.sqd3 or "", fb.sqd4 or "",
                fb.sqd5 or "", fb.sqd6 or "", fb.sqd7 or "", fb.sqd8 or "", fb.sqd9 or "",
                fb.cc1 or "", fb.cc2 or "", fb.cc3 or "",
                round(fb.average_score, 2),
                fb.suggestions or "",
                formatted_date
            ])
        return response

    # Compute average_rating for template display
    for fb in feedback_list:
        scores = [fb.sqd1, fb.sqd2, fb.sqd3, fb.sqd4, fb.sqd5, fb.sqd6, fb.sqd7, fb.sqd8, fb.sqd9]
        valid_scores = [s for s in scores if s is not None]
        fb.average_rating = round(sum(valid_scores)/len(valid_scores), 2) if valid_scores else 0

    return render(
        request,
        "gso_office/feedbacks/feedback_reports.html",
        {
            "feedback_list": feedback_list,
            "units": units,
            "selected_unit": unit_id,
            "selected_month": month,
            "start_date": start_date,
            "end_date": end_date
        }
    )





# GSO Analytics View
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from apps.gso_requests.models import ServiceRequest as Request
from apps.gso_inventory.models import InventoryItem as Material

@login_required
def gso_analytics(request):
    # ===== REQUEST ANALYTICS =====
    total_requests = Request.objects.count()
    completed_requests = Request.objects.filter(status='Completed').count()
    pending_requests = Request.objects.filter(status='Pending').count()
    in_progress_requests = Request.objects.filter(status='In Progress').count()

    # ===== INVENTORY ANALYTICS =====
    total_materials = Material.objects.count()
    low_stock_materials = Material.objects.filter(quantity__lte=10).count()  # threshold for low stock
    out_of_stock = Material.objects.filter(quantity=0).count()

    # ===== CONTEXT =====
    context = {
        # Request analytics
        'total_requests': total_requests,
        'completed_requests': completed_requests,
        'pending_requests': pending_requests,
        'in_progress_requests': in_progress_requests,

        # Inventory analytics
        'total_materials': total_materials,
        'low_stock_materials': low_stock_materials,
        'out_of_stock': out_of_stock,
    }

    return render(request, 'gso_office/analytics/gso_analytics.html', context)





# -------------------------------
# Preview Report (IPMT or WAR)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def preview_report(request):
    """
    Unified preview for IPMT or Work Accomplishment Report (WAR) before export.
    Uses the same filters as the modal: unit, month, personnel.
    """
    report_type = request.GET.get("report_type", "ipmt").lower()
    month_filter = request.GET.get("month")
    unit_filter = request.GET.get("unit")
    personnel_names = request.GET.getlist("personnel[]") or []

    if report_type not in ["ipmt", "war"]:
        return HttpResponse("Invalid report type.", status=400)

    # For IPMT, reuse existing preview_ipmt logic
    if report_type == "ipmt":
        # Call your existing IPMT preview logic
        return preview_ipmt(request)

    # For WAR preview
    if report_type == "war":
        # Parse month
        year = month_num = None
        if month_filter and "-" in month_filter:
            try:
                year, month_num = map(int, month_filter.split("-"))
            except ValueError:
                return HttpResponse("Invalid month format. Use YYYY-MM.", status=400)

        unit = Unit.objects.filter(name__iexact=unit_filter).first() if unit_filter else None

        # Fetch WARs filtered by unit, month, personnel
        wars = WorkAccomplishmentReport.objects.select_related("unit").prefetch_related("assigned_personnel").all()

        if unit:
            wars = wars.filter(unit=unit)
        if year and month_num:
            wars = wars.filter(date_started__year=year, date_started__month=month_num)
        if personnel_names:
            user_objs = [get_user_by_identifier(name) for name in personnel_names]
            user_objs = [u for u in user_objs if u]
            if user_objs:
                wars = wars.filter(assigned_personnel__in=user_objs).distinct()

        # Normalize reports
        reports = []
        for war in wars.order_by("-date_started"):
            reports.append({
                "id": war.id,
                "type": "WorkAccomplishmentReport",
                "date": war.date_started,
                "unit": war.unit.name if war.unit else "Unassigned",
                "description": war.description or "",
                "requesting_office": getattr(war, "requesting_office", ""),
                "personnel": [p.get_full_name() for p in war.assigned_personnel.all()],
                "status": war.status,
                "rating": getattr(war, "rating", "Not Rated"),
                "success_indicator": war.success_indicator or None,
                "request": getattr(war, "request", None),
            })

        context = {
            "reports": reports,
            "month_filter": month_filter,
            "unit_filter": unit_filter,
            "personnel_names": personnel_names,
        }

        return render(request, "gso_office/reports/war_preview.html", context)
    


@login_required
@user_passes_test(is_gso_or_director)
def save_war(request):
    """
    Save edited WAR descriptions from preview.
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)
    try:
        data = json.loads(request.body)
        rows = data.get("rows", [])
    except Exception as e:
        return JsonResponse({"error": f"Invalid JSON: {str(e)}"}, status=400)

    for row in rows:
        war_id = row.get("war_id")
        description = row.get("description", "")
        war = WorkAccomplishmentReport.objects.filter(id=war_id).first()
        if war:
            war.description = description
            war.save()

    return JsonResponse({"status": "success"})

@login_required
@user_passes_test(is_gso_or_director)
def generate_war(request):
    import calendar, os, json
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from django.http import HttpResponse, JsonResponse

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    try:
        data = json.loads(request.body)
        month_filter = data.get("month")
        unit_name = data.get("unit")
        report_ids = data.get("report_ids", [])
    except Exception as e:
        return JsonResponse({"error": f"Invalid request data: {str(e)}"}, status=400)

    if not month_filter:
        return JsonResponse({"error": "Month filter required"}, status=400)

    # Parse month filter
    try:
        year, month_num = map(int, month_filter.split("-"))
    except Exception:
        return JsonResponse({"error": "Invalid month format. Use YYYY-MM"}, status=400)

    month_range = f"{calendar.month_name[month_num]} 1 - {calendar.month_name[month_num]} {calendar.monthrange(year, month_num)[1]}"
    unit = Unit.objects.filter(name__iexact=unit_name).first() if unit_name else None

    # Fetch WARs
    wars = WorkAccomplishmentReport.objects.select_related("unit").prefetch_related("assigned_personnel").filter(
        date_started__year=year,
        date_started__month=month_num
    )
    if unit:
        wars = wars.filter(unit=unit)
    if report_ids:
        wars = wars.filter(id__in=report_ids)
    wars = wars.order_by("date_started")

    # Load template
    template_path = os.path.join(settings.BASE_DIR, "static", "excel_file", "samplewar.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # Merge and write month/unit headers
    ws.merge_cells("A7:H7")
    ws.cell(row=7, column=1, value=month_range).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A9:H9")
    ws.cell(row=9, column=1, value=unit.name.upper() if unit else "UNASSIGNED").alignment = Alignment(horizontal="center", vertical="center")

    start_row = 11

    # Helper: adjust row height dynamically
    from openpyxl.utils import get_column_letter
    def adjust_row_height(row, columns):
        max_lines = 1
        for col in columns:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                col_width = ws.column_dimensions[get_column_letter(col)].width or 10
                lines = len(str(cell.value)) / col_width
                max_lines = max(max_lines, int(lines) + 1)
        ws.row_dimensions[row].height = max(15, max_lines * 15)

    # Write WAR rows
    for i, war in enumerate(wars, start=start_row):
        ws.cell(row=i, column=1).value = war.date_started.strftime("%Y-%m-%d") if war.date_started else ""
        ws.cell(row=i, column=2).value = getattr(war, "date_completed", None).strftime("%Y-%m-%d") if getattr(war, "date_completed", None) else ""
        ws.cell(row=i, column=3).value = getattr(war, "project_name", "")
        ws.cell(row=i, column=4).value = getattr(war, "description", "")
        ws.cell(row=i, column=5).value = getattr(war, "requesting_office", "")
        ws.cell(row=i, column=6).value = ", ".join([p.get_full_name() for p in war.assigned_personnel.all()])
        ws.cell(row=i, column=7).value = getattr(war, "status", "")
        ws.cell(row=i, column=8).value = getattr(war, "rating", "")

        # Wrap text for Project Name and Description, center vertically and horizontally
        for col in [3, 4]:
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        # Center all other columns
        for col in [1,2,5,6,7,8]:
            ws.cell(row=i, column=col).alignment = Alignment(vertical="center", horizontal="center")

        # Adjust row height only based on Project Name and Description
        adjust_row_height(i, [3,4])

    # TOTAL row (optional, if you want to keep a total or leave blank)
    total_row = start_row + len(wars)

    # --- Footer: Prepared / Checked / Noted by ---
    sign_row = total_row + 2
    footer_rows = 3
    unit_name_lower = unit_name.lower() if unit_name else ""

    if "electrical" in unit_name_lower:
        prepared_by_text = "Prepared by:\nPepito Nanalo\nAdministrative Assistant II"
        checked_by_text = "Checked by:\nJuan Juanone\nOIC Head, Electrical Services Unit"
    elif "utility" in unit_name_lower:
        prepared_by_text = "Prepared by:\nLuz Bagani\nAdministrative Assistant II"
        checked_by_text = "Checked by:\nManny Santos\nOIC Head, Utility Unit"
    elif "repair" in unit_name_lower or "maintenance" in unit_name_lower:
        prepared_by_text = "Prepared by:\nRepair/Maintenance Person\nPosition"
        checked_by_text = "Checked by:\nRepair/Maintenance Head\nPosition"
    elif "motorpool" in unit_name_lower:
        prepared_by_text = "Prepared by:\nMotorpool Person\nPosition"
        checked_by_text = "Checked by:\nMotorpool Head\nPosition"
    else:
        prepared_by_text = "Prepared by:\nPrepared Person\nPosition"
        checked_by_text = "Checked by:\nChecked Person\nPosition"

    noted_by_text = "Noted by:\nENGR. Juan Tamad\nGSO Director"

    # Merge cells for 3-row footer
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row+footer_rows-1, end_column=3)
    ws.merge_cells(start_row=sign_row, start_column=4, end_row=sign_row+footer_rows-1, end_column=6)
    ws.merge_cells(start_row=sign_row, start_column=7, end_row=sign_row+footer_rows-1, end_column=8)

    ws.cell(row=sign_row, column=1, value=prepared_by_text)
    ws.cell(row=sign_row, column=4, value=checked_by_text)
    ws.cell(row=sign_row, column=7, value=noted_by_text)

    # Wrap text and center vertically & horizontally
    for col in [1,4,7]:
        cell = ws.cell(row=sign_row, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Adjust row heights for footer
    for r in range(sign_row, sign_row + footer_rows):
        ws.row_dimensions[r].height = 20

    # Return Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"WAR_{unit_name}_{month_filter}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



